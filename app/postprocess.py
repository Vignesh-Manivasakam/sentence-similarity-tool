import pandas as pd
import streamlit as st
import plotly.express as px
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import logging
import difflib
import html
import re

logger = logging.getLogger(__name__)


def df_to_html_table(df, base_data, user_data):
    """Generates a styled HTML table from the results DataFrame with truncation indicators and colored LLM relationships."""
    html_str = '<table class="results-table">'
    
    all_columns = df.columns.tolist()
    
    query_cols = sorted([c for c in all_columns if c.startswith('Query_') and c not in ['Query_Sentence_Highlighted', 'Query_Sentence', 'Query_Sentence_Cleaned_text', 'Query_Object_Identifier']])
    matched_cols = sorted([c for c in all_columns if c.startswith('Matched_') and c not in ['Matched_Sentence_Highlighted', 'Matched_Sentence', 'Matched_Sentence_Cleaned_text', 'Matched_Object_Identifier']])
    
    llm_cols = [c for c in all_columns if c in ['Similarity_Score', 'Similarity_Level', 'Remark'] or 'LLM_' in c]

    # Updated visible_columns 
    visible_columns = (
        ['Query_Object_Identifier', 'Query_Sentence_Highlighted'] +
        query_cols +
        ['Matched_Object_Identifier', 'Matched_Sentence_Highlighted'] +
        matched_cols +
        llm_cols  
    )
    visible_columns = [col for col in visible_columns if col in df.columns]

    html_str += '<tr style="background-color: #444; color: white; text-align: left;">'
    for col in visible_columns:
        header_text = col.replace("_Highlighted", "").replace("_", " ").title()
        html_str += f'<th class="sentence-column" style="width: auto;">{header_text}</th>'
    html_str += '</tr>'

    for idx, row in df.iterrows():
        html_str += '<tr>'
        for col in visible_columns:
            text = str(row.get(col, 'N/A'))
            if col == 'Query_Sentence_Highlighted':
                query_id = row['Query_Object_Identifier']
                is_truncated = next((entry['Truncated'] for entry in user_data if entry['Object_Identifier'] == query_id), False)
                if not is_valid_html(text):
                    text = html.escape(text)
                text = f"<span style='color:yellow'>⚠️</span> {text}" if is_truncated else text
                html_str += f'<td class="sentence-column">{text}</td>'
            elif col == 'Matched_Sentence_Highlighted':
                match_id = row['Matched_Object_Identifier']
                is_truncated = next((entry['Truncated'] for entry in base_data if entry['Object_Identifier'] == match_id), False)
                if not is_valid_html(text):
                    text = html.escape(text)
                text = f"<span style='color:yellow'>⚠️</span> {text}" if is_truncated else text
                html_str += f'<td class="sentence-column">{text}</td>'
            elif col == 'Similarity_Level' or col == 'LLM_Similarity_Level':
                # Enhanced color coding for LLM relationships
                color = '#4CAF50' if any(word in text.lower() for word in ['equivalent', 'exact', 'match']) else \
                        '#ff6b6b' if any(word in text.lower() for word in ['contradictory', 'opposite', 'different']) else \
                        '#FFA500' if any(word in text.lower() for word in ['related', 'similar', 'partial', 'most', 'moderate']) else \
                        '#9E9E9E' if any(word in text.lower() for word in ['threshold', 'error', 'n/a']) else 'inherit'
                html_str += f'<td style="color: {color}; font-weight: bold;">{html.escape(text)}</td>'
            elif col == 'Similarity_Score' or col == 'LLM_Similarity_Score':
                # Format Score with appropriate styling
                if text != 'N/A' and text != 'Error':
                    try:
                        score_val = float(text)
                        color = '#4CAF50' if score_val >= 0.8 else '#FFA500' if score_val >= 0.5 else '#ff6b6b'
                        html_str += f'<td style="color: {color}; font-weight: bold;">{html.escape(text)}</td>'
                    except (ValueError, TypeError):
                        html_str += f'<td style="color: #9E9E9E;">{html.escape(text)}</td>'
                else:
                    html_str += f'<td style="color: #9E9E9E;">{html.escape(text)}</td>'
            else:
                html_str += f'<td>{html.escape(text)}</td>'
        html_str += '</tr>'
    
    html_str += '</table>'
    return html_str if is_valid_html(html_str) else html.escape(html_str)


def is_valid_html(text):
    """Check if text contains valid HTML without nested tags."""
    try:
        invalid_tag_pattern = r'<[\w\s]*[<>][\w\s]*>'
        return not bool(re.search(invalid_tag_pattern, text))
    except:
        return False


def strip_html_tags(text):
    """Remove HTML tags from text and extract plain text."""
    if text is None:
        return ""
    
    if not isinstance(text, str):
        text = str(text)
    
    # Remove HTML tags
    clean = re.compile('<.*?>')
    text_clean = re.sub(clean, '', text)
    
    # Remove extra whitespace
    text_clean = ' '.join(text_clean.split())
    
    return text_clean


def highlight_word_differences(query_text, matched_text):
    """Highlight word-level differences between two sentences for HTML display."""
    try:
        query_words = query_text.split()
        matched_words = matched_text.split()
        diff = list(difflib.ndiff(query_words, matched_words))
        highlighted_query, highlighted_matched = [], []

        for d in diff:
            if d.startswith('  '):
                word = html.escape(d[2:])
                highlighted_query.append(word)
                highlighted_matched.append(word)
            elif d.startswith('- '):
                word = html.escape(d[2:])
                highlighted_query.append(f'<span style="color:#ff6b6b; font-weight:bold">{word}</span>')
            elif d.startswith('+ '):
                word = html.escape(d[2:])
                highlighted_matched.append(f'<span style="color:#00d4ff; font-weight:bold">{word}</span>')

        return ' '.join(highlighted_query), ' '.join(highlighted_matched)
    except:
        return html.escape(str(query_text)), html.escape(str(matched_text))


def create_text_with_diff_markers(query_text, matched_text, is_query=True):
    """
    Create text with simple markers for differences (Excel-safe).
    Uses [DIFF] markers instead of HTML or rich text objects.
    
    Args:
        query_text: Query sentence text
        matched_text: Matched sentence text
        is_query: True for query sentence, False for matched sentence
    
    Returns:
        String with difference markers
    """
    try:
        query_words = str(query_text).split()
        matched_words = str(matched_text).split()
        diff = list(difflib.ndiff(query_words, matched_words))
        
        result_words = []
        
        for d in diff:
            if d.startswith('  '):
                # Common word
                result_words.append(d[2:])
            elif d.startswith('- ') and is_query:
                # Removed word (only in query)
                result_words.append(f"[REMOVED:{d[2:]}]")
            elif d.startswith('+ ') and not is_query:
                # Added word (only in matched)
                result_words.append(f"[ADDED:{d[2:]}]")
        
        return ' '.join(result_words)
    except Exception as e:
        logger.warning(f"Error creating diff markers: {e}")
        return str(query_text if is_query else matched_text)


def sanitize_cell_value(value):
    """
    Sanitize cell values for Excel compatibility.
    Converts all non-primitive types to strings and removes HTML.
    
    Args:
        value: Any cell value
    
    Returns:
        Excel-compatible value (str, int, float, bool, or None)
    """
    # Handle None
    if value is None or (isinstance(value, str) and value.strip() == ''):
        return ""
    
    # Handle numeric types
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    
    # Handle strings - remove HTML tags
    if isinstance(value, str):
        # Remove HTML tags if present
        if '<' in value and '>' in value:
            return strip_html_tags(value)
        return value
    
    # Handle lists (join them)
    if isinstance(value, (list, tuple)):
        return ', '.join(str(v) for v in value)
    
    # Handle dicts
    if isinstance(value, dict):
        return str(value)
    
    # For any other object type, convert to string and remove HTML
    try:
        str_value = str(value)
        if '<' in str_value and '>' in str_value:
            return strip_html_tags(str_value)
        return str_value
    except Exception as e:
        logger.warning(f"Could not convert value to string: {e}")
        return "[Error: Unable to display]"


def create_highlighted_excel(df, base_data, user_data):
    """
    Generate an Excel file with colored cells and difference markers.
    Uses cell styling instead of rich text to avoid TextBlock errors.
    
    Args:
        df: Results DataFrame
        base_data: Base data entries
        user_data: User data entries
    
    Returns:
        BytesIO object with Excel file data
    """
    try:
        logger.info("Creating Excel file with highlighting...")
        
        # ✅ Sanitize DataFrame first
        df_clean = df.copy()
        for col in df_clean.columns:
            df_clean[col] = df_clean[col].apply(sanitize_cell_value)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Similarity Results"

        # Define column order
        all_cols = df_clean.columns.tolist()
        non_meta_query = ['Query_Object_Identifier', 'Query_Sentence', 'Query_Sentence_Cleaned_text', 'Query_Sentence_Highlighted']
        non_meta_matched = ['Matched_Object_Identifier', 'Matched_Sentence', 'Matched_Sentence_Cleaned_text', 'Matched_Sentence_Highlighted']

        query_meta_cols = sorted([c for c in all_cols if c.startswith('Query_') and c not in non_meta_query])
        matched_meta_cols = sorted([c for c in all_cols if c.startswith('Matched_') and c not in non_meta_matched])

        # Include both FAISS and LLM columns
        score_cols = [c for c in all_cols if 'Similarity_Score' in c or 'Similarity_Level' in c or 'Remark' in c or 'LLM_' in c]

        final_headers_ordered = (
            ['Query_Object_Identifier', 'Query_Sentence'] +
            query_meta_cols +
            ['Matched_Object_Identifier', 'Matched_Sentence'] +
            matched_meta_cols +
            score_cols
        )
        headers_to_write = [h for h in final_headers_ordered if h in all_cols]

        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        warning_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        removed_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        added_fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
        border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        # Write headers
        for col_idx, header in enumerate(headers_to_write, 1):
            cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").title())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # Write data rows
        for df_idx, row in df_clean.iterrows():
            excel_row = ws.max_row + 1
            
            for col_idx, col_name in enumerate(headers_to_write, 1):
                cell = ws.cell(row=excel_row, column=col_idx)
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                # Handle sentence columns with difference highlighting
                if col_name == 'Query_Sentence':
                    q_text = sanitize_cell_value(row.get('Query_Sentence', ''))
                    m_text = sanitize_cell_value(row.get('Matched_Sentence', ''))
                    q_id = row.get('Query_Object_Identifier')
                    
                    # Check if truncated
                    is_truncated = False
                    if user_data:
                        is_truncated = next((e.get('Truncated', False) for e in user_data if e.get('Object_Identifier') == q_id), False)
                    
                    # Create text with markers
                    marked_text = create_text_with_diff_markers(q_text, m_text, is_query=True)
                    
                    # Add truncation warning
                    if is_truncated:
                        cell.value = f"⚠️ TRUNCATED: {marked_text}"
                        cell.fill = warning_fill
                    else:
                        cell.value = marked_text
                    
                    # Highlight if there are differences
                    if '[REMOVED:' in marked_text:
                        cell.fill = removed_fill
                
                elif col_name == 'Matched_Sentence':
                    q_text = sanitize_cell_value(row.get('Query_Sentence', ''))
                    m_text = sanitize_cell_value(row.get('Matched_Sentence', ''))
                    m_id = row.get('Matched_Object_Identifier')
                    
                    # Check if truncated
                    is_truncated = False
                    if base_data:
                        is_truncated = next((e.get('Truncated', False) for e in base_data if e.get('Object_Identifier') == m_id), False)
                    
                    # Create text with markers
                    marked_text = create_text_with_diff_markers(q_text, m_text, is_query=False)
                    
                    # Add truncation warning
                    if is_truncated:
                        cell.value = f"⚠️ TRUNCATED: {marked_text}"
                        cell.fill = warning_fill
                    else:
                        cell.value = marked_text
                    
                    # Highlight if there are differences
                    if '[ADDED:' in marked_text:
                        cell.fill = added_fill
                
                else:
                    # Regular columns
                    value = sanitize_cell_value(row.get(col_name))
                    cell.value = value if value != "" else "N/A"
                    
                    # Color code similarity levels
                    if 'Similarity_Level' in col_name:
                        if value and isinstance(value, str):
                            if 'exact' in value.lower() or 'match' in value.lower():
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            elif 'most' in value.lower() or 'similar' in value.lower():
                                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                            elif 'moderate' in value.lower():
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Auto-adjust column widths
        for col_idx in range(1, len(headers_to_write) + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for cell in ws[column_letter]:
                try:
                    if cell.value:
                        # Limit calculation to first 100 chars
                        cell_len = len(str(cell.value)[:100])
                        if cell_len > max_length:
                            max_length = cell_len
                except:
                    pass
            
            adjusted_width = min(max(max_length + 2, 10), 60)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze top row
        ws.freeze_panes = 'A2'

        # Add summary sheet
        try:
            summary_ws = wb.create_sheet("Summary")
            
            # Calculate statistics
            total_results = len(df_clean)
            exact_matches = len(df_clean[df_clean['Similarity_Level'].astype(str).str.contains('Exact', case=False, na=False)]) if 'Similarity_Level' in df_clean.columns else 0
            most_similar = len(df_clean[df_clean['Similarity_Level'].astype(str).str.contains('Most', case=False, na=False)]) if 'Similarity_Level' in df_clean.columns else 0
            moderate = len(df_clean[df_clean['Similarity_Level'].astype(str).str.contains('Moderate', case=False, na=False)]) if 'Similarity_Level' in df_clean.columns else 0
            
            summary_data = [
                ["Metric", "Value"],
                ["Total Results", total_results],
                ["Exact Matches", exact_matches],
                ["Most Similar", most_similar],
                ["Moderately Similar", moderate],
            ]
            
            if base_data and user_data:
                summary_data.extend([
                    ["Base Items", len(base_data)],
                    ["Check Items", len(user_data)],
                ])
            
            for row_idx, row_data in enumerate(summary_data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = summary_ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    if row_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            summary_ws.column_dimensions['A'].width = 25
            summary_ws.column_dimensions['B'].width = 15
        except Exception as e:
            logger.warning(f"Could not create summary sheet: {e}")

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Excel file created successfully with {len(df_clean)} rows")
        return output.getvalue()

    except Exception as e:
        logger.error(f"Error creating highlighted Excel: {e}", exc_info=True)
        
        # ✅ Fallback: Create simple Excel
        try:
            logger.info("Attempting fallback: creating simple Excel...")
            df_simple = df.copy()
            
            # Remove highlighted columns and sanitize
            cols_to_drop = [c for c in df_simple.columns if 'Highlighted' in c]
            df_simple = df_simple.drop(columns=cols_to_drop, errors='ignore')
            
            for col in df_simple.columns:
                df_simple[col] = df_simple[col].apply(sanitize_cell_value)
            
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_simple.to_excel(writer, index=False, sheet_name='Results')
            output.seek(0)
            
            logger.info("Fallback Excel created successfully")
            return output.getvalue()
        except Exception as fallback_error:
            logger.error(f"Fallback Excel creation failed: {fallback_error}")
            raise RuntimeError(f"Could not create Excel file: {str(e)}")


def display_summary(results):
    """Display comprehensive summary statistics with visualizations for both FAISS and LLM analysis."""
    try:
        df = pd.DataFrame(results)
        if df.empty:
            st.warning("No results to display summary for.")
            return
        
        num_queries = len(df['Query_Object_Identifier'].unique()) if 'Query_Object_Identifier' in df.columns else len(df)
        
        # Check which columns are available
        has_llm = 'LLM_Similarity_Level' in df.columns
        has_faiss = 'Similarity_Level' in df.columns
        
        st.markdown('<div class="summary-card">', unsafe_allow_html=True)
        st.markdown("### 📊 Analysis Summary")
        
        # Basic statistics
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric("Total Queries", num_queries)
        with col2:
            st.metric("Total Results", len(df))
        with col3:
            if has_llm and 'LLM_Similarity_Score' in df.columns:
                numeric_scores = pd.to_numeric(df['LLM_Similarity_Score'], errors='coerce').dropna()
                avg_score = numeric_scores.mean() if len(numeric_scores) > 0 else 0
                st.metric("Avg LLM Score", f"{avg_score:.4f}")
            elif has_faiss and 'Similarity_Score' in df.columns:
                numeric_scores = pd.to_numeric(df['Similarity_Score'], errors='coerce').dropna()
                avg_score = numeric_scores.mean() if len(numeric_scores) > 0 else 0
                st.metric("Avg FAISS Score", f"{avg_score:.4f}")
        
        # Create tabs for different views
        if has_llm and has_faiss:
            tab1, tab2, tab3 = st.tabs(["📈 LLM Analysis", "🔍 FAISS Similarity", "📋 Combined View"])
        elif has_llm:
            tab1, tab3 = st.tabs(["📈 LLM Analysis", "📋 Details"]), None
            tab2 = None
        elif has_faiss:
            tab2, tab3 = st.tabs(["🔍 FAISS Similarity", "📋 Details"]), None
            tab1 = None
        else:
            st.warning("No similarity analysis columns found.")
            st.markdown('</div>', unsafe_allow_html=True)
            return
        
        # ========== LLM ANALYSIS TAB ==========
        if has_llm and tab1:
            with tab1:
                st.markdown("#### 🧠 LLM Relationship Analysis")
                
                # LLM Relationship distribution
                if 'LLM_Similarity_Level' in df.columns:
                    llm_relationship_counts = df['LLM_Similarity_Level'].value_counts().to_dict()
                    
                    if llm_relationship_counts:
                        relationship_df = pd.DataFrame.from_dict(
                            llm_relationship_counts, 
                            orient='index', 
                            columns=['Count']
                        ).reset_index().rename(columns={'index': 'Relationship'})
                        
                        # Enhanced color mapping
                        def get_relationship_color(rel):
                            rel_lower = str(rel).lower()
                            if any(word in rel_lower for word in ['exact', 'match', 'equivalent']):
                                return '#4CAF50'  # Green
                            elif any(word in rel_lower for word in ['most', 'similar', 'related']):
                                return '#FFA500'  # Orange
                            elif any(word in rel_lower for word in ['moderate', 'partial']):
                                return '#FFD700'  # Gold
                            elif any(word in rel_lower for word in ['different', 'contradictory', 'opposite']):
                                return '#ff6b6b'  # Red
                            elif any(word in rel_lower for word in ['threshold', 'error', 'n/a']):
                                return '#9E9E9E'  # Gray
                            else:
                                return '#00d4ff'  # Blue (default)
                        
                        colors = [get_relationship_color(rel) for rel in relationship_df['Relationship']]
                        
                        # Create bar chart
                        fig = px.bar(
                            relationship_df,
                            x='Relationship',
                            y='Count',
                            title="Distribution of LLM Relationship Analysis",
                            color='Relationship',
                            color_discrete_sequence=colors,
                            text='Count'
                        )
                        fig.update_traces(
                            textposition='outside',
                            textfont=dict(color='white', size=14, family='Arial Black')
                        )
                        fig.update_layout(
                            paper_bgcolor="#1e1e2f",
                            plot_bgcolor="#2a2a3c",
                            font=dict(color="white"),
                            xaxis_title="Relationship Type",
                            yaxis_title="Count",
                            showlegend=False,
                            xaxis={'tickangle': -45},
                            height=400
                        )
                        st.plotly_chart(fig, use_container_width=True)
                        
                        # Show statistics table
                        st.markdown("**Breakdown by Relationship:**")
                        for _, row in relationship_df.iterrows():
                            percentage = (row['Count'] / len(df)) * 100
                            st.write(f"- **{row['Relationship']}**: {row['Count']} ({percentage:.1f}%)")
                
                # LLM Score distribution
                if 'LLM_Similarity_Score' in df.columns:
                    st.markdown("#### 📊 LLM Score Distribution")
                    numeric_scores = pd.to_numeric(df['LLM_Similarity_Score'], errors='coerce').dropna()
                    
                    if len(numeric_scores) > 0:
                        fig_hist = px.histogram(
                            numeric_scores,
                            nbins=20,
                            title="LLM Similarity Score Distribution",
                            labels={'value': 'Similarity Score', 'count': 'Frequency'},
                            color_discrete_sequence=['#00d4ff']
                        )
                        fig_hist.update_layout(
                            paper_bgcolor="#1e1e2f",
                            plot_bgcolor="#2a2a3c",
                            font=dict(color="white"),
                            showlegend=False,
                            height=350
                        )
                        st.plotly_chart(fig_hist, use_container_width=True)
                        
                        # Score statistics
                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            st.metric("Min Score", f"{numeric_scores.min():.4f}")
                        with col2:
                            st.metric("Max Score", f"{numeric_scores.max():.4f}")
                        with col3:
                            st.metric("Mean Score", f"{numeric_scores.mean():.4f}")
                        with col4:
                            st.metric("Median Score", f"{numeric_scores.median():.4f}")
                
                # Remarks analysis
                if 'LLM_Remark' in df.columns:
                    non_empty_remarks = df[
                        df['LLM_Remark'].notna() & 
                        (df['LLM_Remark'] != '-') & 
                        (df['LLM_Remark'] != '') &
                        (df['LLM_Remark'] != 'N/A')
                    ]
                    st.write(f"**Detailed Analyses with Remarks**: {len(non_empty_remarks)} / {len(df)}")
        
        # ========== FAISS SIMILARITY TAB ==========
        if has_faiss and tab2:
            with tab2:
                st.markdown("#### 🔍 FAISS Similarity Analysis")
                
                # FAISS Level distribution
                if 'Similarity_Level' in df.columns:
                    faiss_level_counts = df['Similarity_Level'].value_counts().to_dict()
                    
                    if faiss_level_counts:
                        level_df = pd.DataFrame.from_dict(
                            faiss_level_counts,
                            orient='index',
                            columns=['Count']
                        ).reset_index().rename(columns={'index': 'Level'})
                        
                        # Color mapping for FAISS levels
                        faiss_colors = {
                            'Exact Match': '#4CAF50',
                            'Most Similar': '#FFA500',
                            'Moderately Similar': '#FFD700',
                            'No Match': '#ff6b6b'
                        }
                        colors = [faiss_colors.get(level, '#00d4ff') for level in level_df['Level']]
                        
                        fig = px.bar(
                            level_df,
                            x='Level',
                            y='Count',
                            title="Distribution of FAISS Similarity Levels",
                            color='Level',
                            color_discrete_sequence=colors,
                            text='Count'
                        )
                        fig.update_traces(
                            textposition='outside',
                            textfont=dict(color='white', size=14, family='Arial Black')
                        )
                        fig.update_layout(
                            paper_bgcolor="#1e1e2f",
                            plot_bgcolor="#2a2a3c",
                            font=dict(color="white"),
                            xaxis_title="Similarity Level",
                            yaxis_title="Count",
                            showlegend=False,
                            height=400
                        )
                        st.plotly_chart(fig, use_container_width=True)
                        
                        # Statistics
                        st.markdown("**Breakdown by Level:**")
                        for _, row in level_df.iterrows():
                            percentage = (row['Count'] / len(df)) * 100
                            st.write(f"- **{row['Level']}**: {row['Count']} ({percentage:.1f}%)")
                
                # FAISS Score distribution
                if 'Similarity_Score' in df.columns:
                    st.markdown("#### 📊 FAISS Score Distribution")
                    numeric_scores = pd.to_numeric(df['Similarity_Score'], errors='coerce').dropna()
                    
                    if len(numeric_scores) > 0:
                        fig_hist = px.histogram(
                            numeric_scores,
                            nbins=20,
                            title="FAISS Similarity Score Distribution",
                            labels={'value': 'Similarity Score', 'count': 'Frequency'},
                            color_discrete_sequence=['#FFD700']
                        )
                        fig_hist.update_layout(
                            paper_bgcolor="#1e1e2f",
                            plot_bgcolor="#2a2a3c",
                            font=dict(color="white"),
                            showlegend=False,
                            height=350
                        )
                        st.plotly_chart(fig_hist, use_container_width=True)
        
        # ========== COMBINED VIEW TAB ==========
        if tab3:
            with tab3:
                st.markdown("#### 📋 Detailed Statistics")
                
                # Create comparison table if both analyses exist
                if has_llm and has_faiss:
                    comparison_data = {
                        'Metric': [],
                        'FAISS Analysis': [],
                        'LLM Analysis': []
                    }
                    
                    # Total analyzed
                    comparison_data['Metric'].append('Total Analyzed')
                    comparison_data['FAISS Analysis'].append(len(df))
                    comparison_data['LLM Analysis'].append(len(df))
                    
                    # Average scores
                    if 'Similarity_Score' in df.columns:
                        faiss_avg = pd.to_numeric(df['Similarity_Score'], errors='coerce').dropna().mean()
                        comparison_data['Metric'].append('Average Score')
                        comparison_data['FAISS Analysis'].append(f"{faiss_avg:.4f}")
                    else:
                        comparison_data['FAISS Analysis'].append('N/A')
                    
                    if 'LLM_Similarity_Score' in df.columns:
                        llm_avg = pd.to_numeric(df['LLM_Similarity_Score'], errors='coerce').dropna().mean()
                        comparison_data['LLM Analysis'].append(f"{llm_avg:.4f}")
                    else:
                        comparison_data['LLM Analysis'].append('N/A')
                    
                    # High quality matches
                    if 'Similarity_Level' in df.columns:
                        high_quality_faiss = len(df[df['Similarity_Level'].str.contains('Exact|Most', case=False, na=False)])
                        comparison_data['Metric'].append('High Quality Matches')
                        comparison_data['FAISS Analysis'].append(high_quality_faiss)
                    
                    if 'LLM_Similarity_Level' in df.columns:
                        high_quality_llm = len(df[df['LLM_Similarity_Level'].str.contains('Exact|Equivalent|Similar', case=False, na=False)])
                        comparison_data['LLM Analysis'].append(high_quality_llm)
                    
                    comparison_df = pd.DataFrame(comparison_data)
                    st.table(comparison_df)
                
                # Token usage if available
                if 'tokens_used' in st.session_state:
                    st.markdown("#### 🔢 LLM Token Usage")
                    tokens = st.session_state['tokens_used']
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Prompt Tokens", f"{tokens.get('prompt_tokens', 0):,}")
                    with col2:
                        st.metric("Completion Tokens", f"{tokens.get('completion_tokens', 0):,}")
                    with col3:
                        total_tokens = tokens.get('prompt_tokens', 0) + tokens.get('completion_tokens', 0)
                        st.metric("Total Tokens", f"{total_tokens:,}")
        
        st.markdown('</div>', unsafe_allow_html=True)
        
    except Exception as e:
        logger.error(f"Error displaying summary: {str(e)}", exc_info=True)
        st.error(f"Error displaying summary: {str(e)}")
        st.info("💡 Try refreshing the page or check the data format.")