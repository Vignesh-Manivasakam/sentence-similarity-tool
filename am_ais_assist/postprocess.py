"""
postprocess.py — HTML table rendering, Excel export, and summary display.

Fix from original:
- Line 70 had `elif col == 'Similarity Score':` (space instead of underscore).
  The actual column name is 'Similarity_Score'. This meant score cells were
  never colour-coded — they always fell through to the plain else branch.
  Fixed to `elif col == 'Similarity_Score':`.

Phase 8 addition:
- create_highlighted_excel() accepts optional include_review_status parameter.
  When True the Review_Status column (pre-added to df by app.py before calling
  this function) is appended to the Excel output with colour-coded fills:
    OK      → light green  (C6EFCE)
    Not OK  → light red    (FFC7CE)
    Pending → no fill
"""

import difflib
import html
import io
import logging
import re

import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import PatternFill

try:
    import bleach
    _BLEACH_AVAILABLE = True
except ImportError:
    _BLEACH_AVAILABLE = False

from am_ais_assist.config import HIGH_CONFIDENCE_THRESHOLD

# H-1: allow-list for HTML tags/attrs emitted by highlight_word_differences()
_ALLOWED_TAGS = {"span", "b", "i", "em", "strong"}
_ALLOWED_ATTRS: dict = {"span": ["style"], "b": [], "i": [], "em": [], "strong": []}


def sanitize_html_cell(text: str) -> str:
    """H-1: Strip all HTML except the known-safe diff-highlight spans.

    Falls back to full html.escape() when bleach is not installed.
    """
    if not _BLEACH_AVAILABLE:
        return html.escape(text)
    return bleach.clean(text, tags=_ALLOWED_TAGS, attributes=_ALLOWED_ATTRS, strip=True,)


# ---------------------------------------------------------------------------
# Colour constants — single source of truth for both HTML table and summary
# ---------------------------------------------------------------------------
_COLOUR_EQUIVALENT = "#4CAF50"  # green
_COLOUR_RELATED = "#FFA500"  # orange
_COLOUR_CONTRADICTORY = "#ff6b6b"  # red
_COLOUR_NEUTRAL = "#9E9E9E"  # grey
_COLOUR_NEW_REQ = "#1565C0"  # blue — new requirement
_COLOUR_DELETED = "#7B1FA2"  # purple — deleted requirement
_COLOUR_OK = "#4CAF50"       # green — review OK
_COLOUR_NOT_OK = "#ff6b6b"   # red — review Not OK
_COLOUR_PENDING = "#9E9E9E"  # grey — review Pending
_SCORE_RELATED_THRESHOLD: float = 0.50

# Excel fill colours for Review_Status column
_FILL_OK = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FILL_NOT_OK = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_FILL_PENDING = PatternFill(fill_type=None)  # no fill


def _level_colour(text: str) -> str:
    """Map a Similarity_Level string to its display colour."""
    t = text.lower()
    if "new requirement" in t:
        return _COLOUR_NEW_REQ
    if "deleted" in t:
        return _COLOUR_DELETED
    if any(w in t for w in ("equivalent", "exact", "match")):
        return _COLOUR_EQUIVALENT
    if any(w in t for w in ("contradictory", "opposite", "different")):
        return _COLOUR_CONTRADICTORY
    if any(w in t for w in ("related", "similar", "partial")):
        return _COLOUR_RELATED
    if any(w in t for w in ("threshold", "error", "n/a", "below")):
        return _COLOUR_NEUTRAL
    return "inherit"


def _review_colour(verdict: str) -> str:
    """Map a Review_Status verdict to its HTML display colour."""
    v = verdict.lower()
    if v == "ok":
        return _COLOUR_OK
    if "not" in v:
        return _COLOUR_NOT_OK
    return _COLOUR_PENDING


def _score_colour(score_val: float) -> str:
    """Map a numeric similarity score to its display colour."""
    if score_val >= HIGH_CONFIDENCE_THRESHOLD:
        return _COLOUR_EQUIVALENT
    if score_val >= _SCORE_RELATED_THRESHOLD:
        return _COLOUR_RELATED
    return _COLOUR_CONTRADICTORY


# ---------------------------------------------------------------------------
# HTML table
# ---------------------------------------------------------------------------


def df_to_html_table(df: pd.DataFrame, base_data: list, user_data: list) -> str:  # noqa: PLR0912
    """
    Generate a styled HTML table from the results DataFrame.

    Columns shown:
      Query identifier → Query sentence (highlighted) → query meta columns
      → Matched identifier → Matched sentence (highlighted) → matched meta columns
      → Similarity_Score, Similarity_Level, Remark
      → Review_Status (if present in df)
    """
    all_columns = df.columns.tolist()

    _query_exclude = {
        "Query_Sentence_Highlighted",
        "Query_Sentence",
        "Query_Sentence_Cleaned_text",
        "Query_Object_Identifier",
    }
    _matched_exclude = {
        "Matched_Sentence_Highlighted",
        "Matched_Sentence",
        "Matched_Sentence_Cleaned_text",
        "Matched_Object_Identifier",
    }

    query_meta_cols = sorted(
        c for c in all_columns if c.startswith("Query_") and c not in _query_exclude
    )
    matched_meta_cols = sorted(
        c for c in all_columns if c.startswith("Matched_") and c not in _matched_exclude
    )
    llm_cols = [c for c in all_columns if c in ("Similarity_Score", "Similarity_Level", "Remark")]

    visible_columns = [
        "Query_Object_Identifier",
        "Query_Sentence_Highlighted",
        *query_meta_cols,
        "Matched_Object_Identifier",
        "Matched_Sentence_Highlighted",
        *matched_meta_cols,
        *llm_cols,
    ]
    # Include Review_Status if it was added by app.py (Phase 6)
    if "Review_Status" in all_columns:
        visible_columns.append("Review_Status")

    visible_columns = [c for c in visible_columns if c in df.columns]

    # Build lookup dicts for O(1) truncation checks instead of O(n) per cell
    user_truncated = {e["Object_Identifier"]: e.get("Truncated", False) for e in user_data}
    base_truncated = {e["Object_Identifier"]: e.get("Truncated", False) for e in base_data}

    rows = ['<table class="results-table">']

    # Header row
    rows.append('<tr style="background-color: #444; color: white; text-align: left;">')
    for col in visible_columns:
        header = col.replace("_Highlighted", "").replace("_", " ").title()
        rows.append(f'<th class="sentence-column" style="width: auto;">{header}</th>')
    rows.append("</tr>")

    # Data rows
    for _, row in df.iterrows():
        rows.append("<tr>")
        for col in visible_columns:
            cell_text = str(row.get(col, "N/A"))

            if col == "Query_Sentence_Highlighted":
                query_id = row["Query_Object_Identifier"]
                cell_text = sanitize_html_cell(cell_text)
                if user_truncated.get(query_id):
                    cell_text = f"<span style='color:yellow'>⚠️</span> {cell_text}"
                rows.append(f'<td class="sentence-column">{cell_text}</td>')

            elif col == "Matched_Sentence_Highlighted":
                match_id = row["Matched_Object_Identifier"]
                cell_text = sanitize_html_cell(cell_text)
                if base_truncated.get(match_id):
                    cell_text = f"<span style='color:yellow'>⚠️</span> {cell_text}"
                rows.append(f'<td class="sentence-column">{cell_text}</td>')

            elif col == "Similarity_Level":
                colour = _level_colour(cell_text)
                rows.append(
                    f'<td style="color: {colour}; font-weight: bold;">'
                    f"{html.escape(cell_text)}</td>"
                )

            elif col == "Similarity_Score":
                if cell_text not in ("N/A", "Error", "None"):
                    try:
                        score_val = float(cell_text)
                        colour = _score_colour(score_val)
                        rows.append(
                            f'<td style="color: {colour}; font-weight: bold;">'
                            f"{html.escape(cell_text)}</td>"
                        )
                    except (ValueError, TypeError):
                        rows.append(
                            f'<td style="color: {_COLOUR_NEUTRAL};">'
                            f"{html.escape(cell_text)}</td>"
                        )
                else:
                    rows.append(
                        f'<td style="color: {_COLOUR_NEUTRAL};">' f"{html.escape(cell_text)}</td>"
                    )

            elif col == "Review_Status":
                colour = _review_colour(cell_text)
                rows.append(
                    f'<td style="color: {colour}; font-weight: bold;">'
                    f"{html.escape(cell_text)}</td>"
                )

            else:
                rows.append(f"<td>{html.escape(cell_text)}</td>")

        rows.append("</tr>")

    rows.append("</table>")
    result = "\n".join(rows)
    return result if is_valid_html(result) else html.escape(result)


# ---------------------------------------------------------------------------
# HTML validation helper
# ---------------------------------------------------------------------------


def is_valid_html(text: str) -> bool:
    """Lightweight check — rejects malformed nested angle brackets."""
    try:
        return not bool(re.search(r"<[\w\s]*[<>][\w\s]*>", text))
    except Exception:
        return False


# ---------------------------------------------------------------------------
# Word-level diff highlighting
# ---------------------------------------------------------------------------


def highlight_word_differences(query_text: str, matched_text: str) -> tuple[str, str]:
    """
    Return (highlighted_query, highlighted_matched) HTML strings
    where differing words are coloured.
    """
    try:
        q_words = str(query_text).split()
        m_words = str(matched_text).split()
        diff = list(difflib.ndiff(q_words, m_words))

        highlighted_query: list[str] = []
        highlighted_matched: list[str] = []

        for token in diff:
            if token.startswith("  "):
                word = html.escape(token[2:])
                highlighted_query.append(word)
                highlighted_matched.append(word)
            elif token.startswith("- "):
                word = html.escape(token[2:])
                highlighted_query.append(
                    f'<span style="color:#ff6b6b; font-weight:bold">{word}</span>'
                )
            elif token.startswith("+ "):
                word = html.escape(token[2:])
                highlighted_matched.append(
                    f'<span style="color:#00d4ff; font-weight:bold">{word}</span>'
                )

        return " ".join(highlighted_query), " ".join(highlighted_matched)

    except Exception:
        return html.escape(str(query_text)), html.escape(str(matched_text))


# ---------------------------------------------------------------------------
# Excel export
# Phase 8: Added include_review_status parameter.
# ---------------------------------------------------------------------------


def create_highlighted_excel(  # noqa: PLR0912, PLR0915
    df: pd.DataFrame,
    base_data: list,
    user_data: list,
    include_review_status: bool = False,
) -> bytes:
    """
    Generate an Excel file with word-level diff highlighting.
    Deleted words in red, added words in blue, truncation warnings in yellow.

    Phase 8 addition:
        include_review_status (bool): When True and the df contains a
        'Review_Status' column, it is appended as the final column with
        colour-coded cell fills (green=OK, red=Not OK, no fill=Pending).
        The df must already contain the Review_Status column — this
        function does not look it up from session state.
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "LLM Analysis Results"

        all_cols = df.columns.tolist()
        _non_meta_query = {
            "Query_Object_Identifier",
            "Query_Sentence",
            "Query_Sentence_Cleaned_text",
            "Query_Sentence_Highlighted",
        }
        _non_meta_matched = {
            "Matched_Object_Identifier",
            "Matched_Sentence",
            "Matched_Sentence_Cleaned_text",
            "Matched_Sentence_Highlighted",
        }

        query_meta_cols = sorted(
            c for c in all_cols if c.startswith("Query_") and c not in _non_meta_query
        )
        matched_meta_cols = sorted(
            c for c in all_cols if c.startswith("Matched_") and c not in _non_meta_matched
        )

        final_headers_ordered = (
            ["Query_Object_Identifier", "Query_Sentence"]
            + query_meta_cols
            + ["Matched_Object_Identifier", "Matched_Sentence"]
            + matched_meta_cols
            + ["Similarity_Score", "Similarity_Level", "Remark"]
        )

        # Phase 8: include Review_Status as the last column when requested
        if include_review_status and "Review_Status" in all_cols:
            final_headers_ordered.append("Review_Status")

        headers_to_write = [h for h in final_headers_ordered if h in all_cols]
        ws.append([h.replace("_", " ").title() for h in headers_to_write])

        # Bold the header row
        from openpyxl.styles import Font as _Font
        for cell in ws[1]:
            cell.font = _Font(bold=True)

        # Build O(1) truncation lookups
        user_truncated = {e["Object_Identifier"]: e.get("Truncated", False) for e in user_data}
        base_truncated = {e["Object_Identifier"]: e.get("Truncated", False) for e in base_data}

        df_to_write = df[headers_to_write]

        # Map column index of Review_Status (1-based) for fill application
        review_status_col_idx: int | None = None
        if include_review_status and "Review_Status" in headers_to_write:
            review_status_col_idx = headers_to_write.index("Review_Status") + 1

        for _, row in df_to_write.iterrows():
            row_idx = ws.max_row + 1
            q_text = str(row["Query_Sentence"])
            m_text = str(row["Matched_Sentence"])

            # Skip ndiff highlighting for placeholder/gap-analysis rows
            from am_ais_assist.config import PLACEHOLDER_DELETED_QUERY, PLACEHOLDER_NO_MATCH

            _is_placeholder = m_text == PLACEHOLDER_NO_MATCH or q_text == PLACEHOLDER_DELETED_QUERY
            diff = [] if _is_placeholder else list(difflib.ndiff(q_text.split(), m_text.split()))

            for col_idx, col_name in enumerate(headers_to_write, start=1):
                if col_name in ("Query_Sentence", "Matched_Sentence"):
                    blocks: list = []
                    q_id = row["Query_Object_Identifier"]
                    m_id = row["Matched_Object_Identifier"]

                    # For placeholder rows skip rich-text diff and write plain
                    if _is_placeholder:
                        ws.cell(row=row_idx, column=col_idx, value=str(row[col_name]))
                        continue

                    if col_name == "Query_Sentence":
                        if user_truncated.get(q_id):
                            blocks.append(TextBlock(InlineFont(color="FFFF00", b=True), "⚠️ "))
                        for token in diff:
                            if token.startswith("- "):
                                blocks.append(
                                    TextBlock(InlineFont(color="FF0000", b=True), token[2:] + " ")
                                )
                            elif token.startswith("  "):
                                blocks.append(
                                    TextBlock(InlineFont(color="000000"), token[2:] + " ")
                                )

                    else:  # Matched_Sentence
                        if base_truncated.get(m_id):
                            blocks.append(TextBlock(InlineFont(color="FFFF00", b=True), "⚠️ "))
                        for token in diff:
                            if token.startswith("+ "):
                                blocks.append(
                                    TextBlock(InlineFont(color="0000FF", b=True), token[2:] + " ")
                                )
                            elif token.startswith("  "):
                                blocks.append(
                                    TextBlock(InlineFont(color="000000"), token[2:] + " ")
                                )

                    ws.cell(row=row_idx, column=col_idx).value = (
                        CellRichText(blocks) if blocks else row[col_name]
                    )

                elif col_name == "Review_Status" and review_status_col_idx is not None:
                    # Phase 8: write verdict and apply colour fill
                    verdict = str(row.get(col_name, "Pending"))
                    cell = ws.cell(row=row_idx, column=col_idx, value=verdict)
                    v_lower = verdict.lower()
                    if v_lower == "ok":
                        cell.fill = _FILL_OK
                    elif "not" in v_lower:
                        cell.fill = _FILL_NOT_OK
                    # else: Pending — no fill

                else:
                    value = row[col_name]
                    ws.cell(
                        row=row_idx,
                        column=col_idx,
                        value=str(value) if pd.notna(value) else "N/A",
                    )

        # Auto-fit column widths
        for col in ws.columns:
            col_letter = col[0].column_letter
            max_len = 0
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception as exc:  # noqa: BLE001
                    logging.debug("Could not measure cell width: %s", exc)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except Exception as exc:
        logging.error("Error creating Excel file: %s", exc)
        raise


# ---------------------------------------------------------------------------
# Summary panel
# ---------------------------------------------------------------------------

# Colour map for the summary bar chart — matches your 3 LLM labels
# plus the pipeline sentinel values (Exact Match, Below Threshold, errors).
_SUMMARY_COLOUR_MAP: dict[str, str] = {
    "Equivalent": _COLOUR_EQUIVALENT,
    "Exact Match": _COLOUR_EQUIVALENT,
    "Related": _COLOUR_RELATED,
    "Contradictory": _COLOUR_CONTRADICTORY,
    "Below Threshold": _COLOUR_NEUTRAL,
    "Analysis Error": _COLOUR_NEUTRAL,
    "Error": _COLOUR_NEUTRAL,
    "N/A": _COLOUR_NEUTRAL,
    "New Requirement": _COLOUR_NEW_REQ,
    "Deleted": _COLOUR_DELETED,
}


def _summary_colour(relationship: str) -> str:
    """Resolve a relationship label to a chart colour."""
    if relationship in _SUMMARY_COLOUR_MAP:
        return _SUMMARY_COLOUR_MAP[relationship]
    return _level_colour(relationship)


def display_summary(results: list) -> None:
    """Render the analysis summary panel with statistics and a bar chart."""
    try:
        df = pd.DataFrame(results)
        if df.empty:
            st.warning("No results to display summary for.")
            return

        num_queries = df["Query_Object_Identifier"].nunique()

        if "Similarity_Score" in df.columns:
            numeric_scores = pd.to_numeric(df["Similarity_Score"], errors="coerce").dropna()
            avg_score = numeric_scores.mean() if len(numeric_scores) > 0 else 0.0

            level_counts = (
                df["Similarity_Level"].value_counts().to_dict()
                if "Similarity_Level" in df.columns
                else {}
            )

            st.markdown('<div class="summary-card">', unsafe_allow_html=True)
            st.markdown("**Analysis Summary**")
            st.write(f"- **Queries Processed**: {num_queries}")
            st.write(f"- **Average Similarity Score**: {avg_score:.4f}")
            st.write(f"- **Total Analyses**: {len(df)}")

            if "Remark" in df.columns:
                meaningful_remarks = df[
                    df["Remark"].notna()
                    & (df["Remark"] != "")
                    & (df["Remark"] != "-")
                    & (df["Remark"] != "Exactly Matched")
                ]
                st.write(f"- **Analyses with Remarks**: {len(meaningful_remarks)}")

            # New/Deleted counts from hierarchical comparison
            new_count = level_counts.get("New Requirement", 0)
            deleted_count = level_counts.get("Deleted", 0)
            if new_count or deleted_count:
                st.write(f"- **New Requirements** (not in base): {new_count}")
                st.write(f"- **Deleted Requirements** (removed): {deleted_count}")

            if level_counts:
                rel_df = (
                    pd.DataFrame.from_dict(level_counts, orient="index", columns=["Count"])
                    .reset_index()
                    .rename(columns={"index": "Relationship"})
                )
                colours = [_summary_colour(r) for r in rel_df["Relationship"]]

                fig = px.bar(
                    rel_df,
                    x="Relationship",
                    y="Count",
                    title="Distribution of Relationship Analysis",
                    color="Relationship",
                    color_discrete_sequence=colours,
                    text="Count",
                )
                fig.update_traces(
                    textposition="inside",
                    textfont=dict(color="white", size=15, family="Arial Black"),
                )
                fig.update_layout(
                    paper_bgcolor="#1e1e2f",
                    plot_bgcolor="#2a2a3c",
                    font=dict(color="white"),
                    xaxis_title="Relationship",
                    yaxis_title="Count",
                    showlegend=False,
                    xaxis={"tickangle": 45},
                )
                st.plotly_chart(fig, use_container_width=True)

        else:
            avg_similarity = (
                pd.to_numeric(df["Similarity_Score"], errors="coerce").mean()
                if "Similarity_Score" in df.columns
                else 0.0
            )
            level_counts = (
                df["Similarity_Level"].value_counts().to_dict()
                if "Similarity_Level" in df.columns
                else {}
            )

            st.markdown('<div class="summary-card">', unsafe_allow_html=True)
            st.markdown("**Similarity Summary** (LLM analysis not available)")
            st.write(f"- **Queries Processed**: {num_queries}")
            st.write(f"- **Average Similarity Score**: {avg_similarity:.4f}")

            if level_counts:
                level_df = (
                    pd.DataFrame.from_dict(level_counts, orient="index", columns=["Count"])
                    .reset_index()
                    .rename(columns={"index": "Category"})
                )
                fig = px.bar(
                    level_df,
                    x="Category",
                    y="Count",
                    title="Distribution of Similarity Levels",
                    color="Category",
                    color_discrete_sequence=["#00d4ff", "#ffaa00", "#00ffaa", "#ff6b6b"] * 2,
                    text="Count",
                )
                fig.update_traces(
                    textposition="inside",
                    textfont=dict(color="brown", size=15, family="Arial Black"),
                )
                fig.update_layout(
                    paper_bgcolor="#1e1e2f",
                    plot_bgcolor="#2a2a3c",
                    font=dict(color="white"),
                    xaxis_title="Category",
                    yaxis_title="Count",
                    showlegend=False,
                )
                st.plotly_chart(fig, use_container_width=True)

        st.markdown("</div>", unsafe_allow_html=True)

    except Exception as exc:
        logging.error("Error displaying summary: %s", exc)
        st.error(f"Error displaying summary: {exc}")
